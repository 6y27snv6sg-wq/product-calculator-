import os
import sys

# تثبيت openpyxl تلقائياً في حال عدم وجودها
try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", "openpyxl"]
    )
    import openpyxl

import datetime
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
import streamlit as st

# 1. إعدادات الصفحة
st.set_page_config(
    page_title="حاسبة تكاليف المنتجات", page_icon="📱", layout="centered"
)

# 2. قاعدة بيانات المشتركين وتاريخ التفعيل
SUBSCRIBERS_DATABASE = {
    "PRO-2026": datetime.date(2026, 8, 30),
    "VIP-9999": datetime.date(2026, 8, 30),
}

st.sidebar.title("🔐 بوابة المشتركين")
user_key = st.sidebar.text_input("أدخل مفتاح الاشتراك الخاص بك:", type="password")

if not user_key:
    st.title("📱 حاسبة تكاليف وهامش ربح المنتجات")
    st.warning(
        "⚠️ يرجى إدخال مفتاح الاشتراك في القائمة الجانبية للوصول إلى الخدمة."
    )
    st.stop()

if user_key not in SUBSCRIBERS_DATABASE:
    st.error("❌ مفتاح الاشتراك غير صحيح.")
    st.stop()

# فحص تاريخ انتهاء الاشتراك تلقائياً (بعد 365 يوم)
activation_date = SUBSCRIBERS_DATABASE[user_key]
expiration_date = activation_date + datetime.timedelta(days=365)
today = datetime.date.today()

if today > expiration_date:
    st.error(
        f"❌ انتهت صلاحية اشتراكك بتاريخ {expiration_date}. يرجى التواصل مع الدعم للتجديد."
    )
    st.stop()

days_left = (expiration_date - today).days

# 3. الواجهة الرئيسية للمشترك
st.title("📱 حاسبة تكاليف المنتجات والربحية")
st.success(f"مرحباً بك! اشتراكك فعال (متبقي {days_left} يوم على الانتهاء).")

st.markdown("---")

st.subheader("📦 1. بيانات المنتج والتكاليف المباشرة")
col1, col2 = st.columns(2)
with col1:
    sku = st.text_input("رمز المنتج (SKU)", value="SKU-001")
    buy_price = st.number_input("تكلفة الشراء (ر.س)", value=45.0, step=1.0)
with col2:
    product_name = st.text_input("اسم المنتج", value="سماعات لاسلكية")
    shipping_price = st.number_input(
        "تكلفة الشحن والجمارك (ر.س)", value=5.0, step=1.0
    )

marketing_cost = st.number_input(
    "تكلفة التسويق للوحدة (ر.س)", value=15.0, step=1.0
)

st.subheader("🏛️ 2. الرسوم والضرائب والمبيعات")
col3, col4, col5 = st.columns(3)
with col3:
    gateway_commission_pct = st.number_input(
        "عمولة البوابة (%)", value=2.5, step=0.1
    )
with col4:
    vat_pct = st.number_input("الضريبة (%)", value=15.0, step=1.0)
with col5:
    target_sell_price = st.number_input(
        "سعر البيع المستهدف (ر.س)", value=120.0, step=5.0
    )

# الحسابات
vat_amount = target_sell_price * (vat_pct / 100.0)
gateway_fee = target_sell_price * (gateway_commission_pct / 100.0)
total_unit_cost = (
    buy_price + shipping_price + marketing_cost + gateway_fee + vat_amount
)
gross_profit = target_sell_price - total_unit_cost
profit_margin_pct = (
    (gross_profit / target_sell_price * 100) if target_sell_price > 0 else 0
)

st.markdown("---")
st.subheader("📊 ملخص البطاقة المالية للمنتج")

kpi1, kpi2, kpi3 = st.columns(3)
kpi1.metric("إجمالي التكلفة الشاملة", f"{total_unit_cost:.2f} ر.س")
kpi2.metric(
    "صافي هامش الربح",
    f"{gross_profit:.2f} ر.س",
    delta=f"{profit_margin_pct:.1f}%",
)
kpi3.metric("سعر البيع المستهدف", f"{target_sell_price:.2f} ر.س")


def create_excel_card():
    wb = Workbook()
    ws = wb.active
    ws.title = f"بطاقة_{sku}"
    ws.views.sheetView[0].showGridLines = True

    fill_header = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    fill_profit = PatternFill(
        start_color="C6EFCE" if gross_profit > 0 else "FFC7CE", fill_type="solid"
    )
    font_header = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_profit = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="006100" if gross_profit > 0 else "9C0006",
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("B2:D2")
    ws["B2"] = f"بطاقة منتج: {product_name}"
    ws["B2"].fill = fill_header
    ws["B2"].font = font_header
    ws["B2"].alignment = align_center

    card_data = [
        ("رمز المنتج (SKU)", sku),
        ("اسم المنتج", product_name),
        ("تكلفة الشراء (ر.س)", f"{buy_price:.2f}"),
        ("تكلفة الشحن والجمارك (ر.س)", f"{shipping_price:.2f}"),
        ("تكلفة التسويق للوحدة (ر.س)", f"{marketing_cost:.2f}"),
        ("نسبة عمولة البوابة", f"{gateway_commission_pct}%"),
        ("عمولة البوابة المقدرة (ر.س)", f"{gateway_fee:.2f}"),
        ("نسبة ضريبة القيمة المضافة", f"{vat_pct}%"),
        ("قيمة الضريبة (ر.س)", f"{vat_amount:.2f}"),
        ("إجمالي تكلفة الوحدة (ر.س)", f"{total_unit_cost:.2f}"),
        ("سعر البيع المستهدف (ر.س)", f"{target_sell_price:.2f}"),
        ("هامش الربح الإجمالي (ر.س)", f"{gross_profit:.2f}"),
        ("نسبة هامش الربح", f"{profit_margin_pct:.1f}%"),
    ]

    for row_idx, (label, val) in enumerate(card_data, start=3):
        ws.cell(row=row_idx, column=2, value=label).font = font_bold
        ws.cell(row=row_idx, column=2).alignment = align_right

        val_cell = ws.cell(row=row_idx, column=3, value=val)
        val_cell.alignment = align_center

        if "هامش الربح" in label:
            val_cell.fill = fill_profit
            val_cell.font = font_profit
        elif label == "إجمالي تكلفة الوحدة (ر.س)":
            val_cell.font = font_bold

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 2

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()
    return temp_file.name


excel_path = create_excel_card()
with open(excel_path, "rb") as file:
    st.download_button(
        label="📥 تحميل بطاقة المنتج (Excel)",
        data=file,
        file_name=f"Product_Card_{sku}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
